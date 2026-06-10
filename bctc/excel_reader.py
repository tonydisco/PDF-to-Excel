# -*- coding: utf-8 -*-
"""
Đọc lại file Excel DO APP NÀY XUẤT (Thông tư 200) -> results = {key: {code: (cur, prior)}}.
Dùng cho tính năng "tải Excel lên để phân tích tài chính".

VALIDATE chặt: chỉ chấp nhận đúng format app tạo — sheet tên chuẩn + header dòng 4
có cột "Mã số". Sai format -> ExcelFormatError (thông điệp tiếng Việt rõ ràng).

Layout mỗi sheet (xem excel_writer._write_sheet):
  A1: tiêu đề (merge) · A2: "Đơn vị tính: VND"
  Dòng 4: ["Chỉ tiêu", "Mã số", "Số cuối năm"/"Năm nay", "Số đầu năm"/"Năm trước"]
  Dòng 5+: A=chỉ tiêu, B=mã số, C=cuối năm/năm nay, D=đầu năm/năm trước
"""
import openpyxl

SHEET_TO_KEY = {
    "bảng cân đối kế toán": "CDKT",
    "kết quả hđkd": "KQHDKD",
    "lưu chuyển tiền tệ": "LCTT",
}
HEADER_ROW = 4
DATA_START = 5


class ExcelFormatError(Exception):
    """File Excel không đúng định dạng app xuất (trả 400 + thông điệp tiếng Việt)."""


def _norm(s):
    return str(s if s is not None else "").strip().lower()


def _to_num(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v) == int(v) else float(v)
    s = str(v).strip()
    if not s or s in ("-", "—", "·"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(",", "").replace(" ", "").replace(" ", "")
    try:
        n = float(s)
    except ValueError:
        return None
    n = -n if neg else n
    return int(n) if n == int(n) else n


def read_results(path):
    """Trả results = {key: {code: (cur, prior)}}. RAISE ExcelFormatError nếu sai format."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise ExcelFormatError(f"Không mở được file Excel: {e}")

    results = {}
    recognized = 0
    try:
        for ws in wb.worksheets:
            key = SHEET_TO_KEY.get(_norm(ws.title))
            if not key:
                continue
            # Xác nhận đúng format: cột B dòng 4 phải là "Mã số".
            if _norm(ws.cell(row=HEADER_ROW, column=2).value) != "mã số":
                continue
            recognized += 1
            vals = {}
            for row in ws.iter_rows(min_row=DATA_START, values_only=True):
                if len(row) < 4:
                    continue
                code = row[1]
                code = str(code).strip() if code is not None else ""
                if not code:
                    continue
                vals[code] = (_to_num(row[2]), _to_num(row[3]))
            if vals:
                results[key] = vals
    finally:
        wb.close()

    if recognized == 0:
        raise ExcelFormatError(
            "File Excel không đúng định dạng app xuất. Cần có ít nhất một sheet "
            "'Bảng cân đối kế toán' / 'Kết quả HĐKD' / 'Lưu chuyển tiền tệ' với cột 'Mã số' ở dòng 4."
        )
    if not results:
        raise ExcelFormatError("File Excel đúng định dạng nhưng không có dòng dữ liệu (Mã số) nào.")
    return results
