# -*- coding: utf-8 -*-
"""Xuất kết quả ra file Excel: mỗi báo cáo = 1 sheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import templates as T

HEADERS = {
    "CDKT":   ("Mã số", "Số cuối năm", "Số đầu năm"),
    "KQHDKD": ("Mã số", "Năm nay", "Năm trước"),
    "LCTT":   ("Mã số", "Năm nay", "Năm trước"),
}
SHEET_NAMES = {
    "CDKT": "Bảng cân đối kế toán",
    "KQHDKD": "Kết quả HĐKD",
    "LCTT": "Lưu chuyển tiền tệ",
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")
FLAG_FILL = PatternFill("solid", fgColor="FFE08A")   # ô nghi ngờ -> tô cam nhạt
NUMFMT = '#,##0;(#,##0)'

# Nhãn cột trong danh sách conflicts -> số cột Excel
_CONFLICT_COL = {"cuối năm/năm nay": 3, "đầu năm/năm trước": 4}


def pick_lctt_template(values):
    gt = T.codes_of(T.LUU_CHUYEN_TIEN_TE_GT)
    tt = T.codes_of(T.LUU_CHUYEN_TIEN_TE_TT)
    have = set(values)
    return T.LUU_CHUYEN_TIEN_TE_TT if len(have & tt) > len(have & gt) else T.LUU_CHUYEN_TIEN_TE_GT


def _template_for(stmt_key, values):
    """Khung template dùng để ghi sheet (LCTT chọn gián tiếp/trực tiếp theo mã)."""
    if stmt_key == "LCTT":
        return pick_lctt_template(values)
    return T.STATEMENTS[stmt_key][1]


def extra_codes(stmt_key, values):
    """Mã trong `values` KHÔNG có trong khung template của stmt_key. §7.6/G2:
    những mã này sẽ BIẾN MẤT khỏi sheet nếu chỉ duyệt template -> phải hiện ra.
    Giữ thứ tự xuất hiện trong `values`."""
    template = _template_for(stmt_key, values)
    tpl_codes = {row[0] for row in template if row[0] is not None}
    return [c for c in values if c not in tpl_codes]


def out_of_framework_warnings(results):
    """Cảnh báo (tiếng Việt) cho mỗi báo cáo có mã ngoài khung. Engine đưa vào
    warnings từng file để đóng G2 (không mất dữ liệu âm thầm)."""
    warns = []
    for key in ("CDKT", "KQHDKD", "LCTT"):
        vals = results.get(key)
        if not vals:
            continue
        extra = extra_codes(key, vals)
        if extra:
            warns.append(
                "%s: %d mã NGOÀI KHUNG được bóc ra (%s) — xem mục 'MÃ NGOÀI "
                "KHUNG' cuối sheet, cần kiểm tra." % (
                    SHEET_NAMES[key], len(extra), ", ".join(extra)))
    return warns


def _write_sheet(ws, stmt_key, title_full, values, flags=frozenset(), unit=None):
    template = _template_for(stmt_key, values)
    h_code, h_cur, h_prior = HEADERS[stmt_key]

    # ---- tiêu đề ----
    ws.merge_cells("A1:D1")
    ws["A1"] = title_full.upper()
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    # §7.3: đơn vị tính phát hiện được (mặc định VND). KHÔNG quy đổi giá trị.
    ws["A2"] = "Đơn vị tính: " + (unit or "VND")
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="right")

    # ---- hàng tiêu đề cột ----
    hr = 4
    cols = ["Chỉ tiêu", h_code, h_cur, h_prior]
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=hr, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # ---- dữ liệu ----
    r = hr + 1
    for code, label, level, kind in template:
        cur = prior = None
        if code in values:
            cur, prior = values[code]
        c_label = ws.cell(row=r, column=1, value=("    " * level) + label)
        c_code = ws.cell(row=r, column=2, value=code if code else "")
        c_cur = ws.cell(row=r, column=3, value=cur)
        c_prior = ws.cell(row=r, column=4, value=prior)

        for cc in (c_label, c_code, c_cur, c_prior):
            cc.border = BORDER
        c_code.alignment = Alignment(horizontal="center")
        c_cur.number_format = NUMFMT
        c_prior.number_format = NUMFMT
        c_cur.alignment = c_prior.alignment = Alignment(horizontal="right")

        bold = kind in ("header", "section", "total")
        if bold:
            f = Font(bold=True)
            for cc in (c_label, c_code, c_cur, c_prior):
                cc.font = f
        if kind in ("header", "section"):
            for cc in (c_label, c_code, c_cur, c_prior):
                cc.fill = SECTION_FILL
        elif kind == "total":
            for cc in (c_label, c_code, c_cur, c_prior):
                cc.fill = TOTAL_FILL
        # tô cam ô "nghi ngờ" (hai lần OCR đọc lệch) — ưu tiên hơn fill nền
        if (code, 3) in flags:
            c_cur.fill = FLAG_FILL
        if (code, 4) in flags:
            c_prior.fill = FLAG_FILL
        r += 1

    # ---- §7.6/G2: MÃ NGOÀI KHUNG (không mất dữ liệu âm thầm) ----
    tpl_codes = {row[0] for row in template if row[0] is not None}
    extra = [c for c in values if c not in tpl_codes]
    if extra:
        r += 1                                  # chừa một dòng trống
        for col in range(1, 5):
            hc = ws.cell(row=r, column=col)
            hc.fill = FLAG_FILL
            hc.border = BORDER
        head = ws.cell(row=r, column=1, value="MÃ NGOÀI KHUNG (cần kiểm tra)")
        head.font = Font(bold=True, color="C00000")
        r += 1
        for code in extra:
            cur, prior = values[code]
            c_label = ws.cell(row=r, column=1,
                              value="(không có trong khung Thông tư 200)")
            c_code = ws.cell(row=r, column=2, value=code)
            c_cur = ws.cell(row=r, column=3, value=cur)
            c_prior = ws.cell(row=r, column=4, value=prior)
            for cc in (c_label, c_code, c_cur, c_prior):
                cc.border = BORDER
                cc.fill = FLAG_FILL
            c_code.alignment = Alignment(horizontal="center")
            c_cur.number_format = NUMFMT
            c_prior.number_format = NUMFMT
            c_cur.alignment = c_prior.alignment = Alignment(horizontal="right")
            r += 1

    # ---- định dạng cột ----
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def _flags_by_key(conflicts):
    """[(key, code, label, v1, v2)] -> {key: {(code, colno)}}"""
    out = {}
    for key, code, label, _v1, _v2 in (conflicts or []):
        colno = _CONFLICT_COL.get(label)
        if colno:
            out.setdefault(key, set()).add((code, colno))
    return out


def build_workbook(company_name, results, conflicts=None, unit=None):
    flags = _flags_by_key(conflicts)
    wb = Workbook()
    wb.remove(wb.active)
    created = 0
    # Chỉ tạo sheet cho báo cáo CÓ bóc được chỉ tiêu; báo cáo trống thì BỎ QUA.
    for key in ("CDKT", "KQHDKD", "LCTT"):
        if not results.get(key):
            continue
        ws = wb.create_sheet(title=SHEET_NAMES[key])
        _write_sheet(ws, key, T.STATEMENTS[key][0], results[key],
                     flags.get(key, frozenset()), unit=unit)
        created += 1
    if created == 0:
        # openpyxl cần tối thiểu 1 sheet -> ghi sheet thông báo thay vì file rỗng.
        ws = wb.create_sheet(title="Không có dữ liệu")
        ws["A1"] = ("Không nhận diện được báo cáo nào theo mẫu Thông tư 200 "
                    "(file có thể theo mẫu khác, ví dụ ngân hàng/TCTD).")
        ws["A1"].font = Font(bold=True, color="C00000")
        ws.column_dimensions["A"].width = 90
    return wb


def save(company_name, results, out_path, conflicts=None, unit=None):
    wb = build_workbook(company_name, results, conflicts, unit=unit)
    wb.save(out_path)
    return out_path
