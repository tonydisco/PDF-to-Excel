# -*- coding: utf-8 -*-
"""
Đọc file Excel "đáp án" (bản xuất trực tiếp từ phần mềm kế toán).

Bốn cạm bẫy đã xác nhận trên corpus thật:
  1. Đuôi file nói dối: 'KQKD.XLS' thực chất là OOXML (bắt đầu bằng 'PK').
     openpyxl từ chối theo ĐUÔI FILE chứ không theo nội dung -> phải nạp qua
     BytesIO sau khi tự dò magic bytes.
  2. Nhãn chỉ tiêu là mojibake TCVN3 ('C«ng ty CP Du LÞch §¨k L¨k'). Không cần
     giải mã: ta chỉ dùng MÃ SỐ + CON SỐ.
  3. Hàng tiêu đề không ở vị trí cố định -> phải dò động.
  4. Cột "năm nay"/"năm trước" không LUÔN LUÔN trái = năm nay: file thật
     'Luu chuyen tien te 2015.xls' (Cty CP DL KS Tháng Mười) có nhãn tiêu đề
     ĐỌC ĐƯỢC 'Kỳ trước' | 'Kỳ này' với cột TRÁI lại là Kỳ trước. Khi nhãn
     còn đọc được thì phải ưu tiên nhãn hơn vị trí, xem _find_columns_fallback
     và statement_detail().
  5. Hàng đánh số '1..5' nằm trên Ô MERGE nên marker '4'/'5' có thể lệch 1
     cột so với cột giá trị thật ('KQKD_2023.xls': marker ở c9/c11, giá trị
     ở c8/c10 -> key toàn (None, None); 'cdkt_6t.xls'/'CDKT Q2.xls': marker
     '5' ở cột rỗng, Kỳ-trước thật ở kề trái -> mất nguyên cột năm trước).
     Cột neo theo hàng đánh số phải được XÁC MINH có ô số thật trên các hàng
     có mã; cột rỗng thì dò 2 cột kề, xem _nudge_empty_value_col.
  6. Bảng 4 cột giá trị theo NHÓM header (Trong kỳ nay/trước + Lũy kế
     nay/trước — 'BTHT ..._KQKD.xls'): "2 cột nhiều số nhất" có thể là 2 cột
     Năm-Trước của 2 nhóm khác nhau (vài dòng chỉ có số năm trước). Khi nhãn
     còn đọc được và có cả cột năm-nay lẫn năm-trước trong các ứng viên thì
     phải ghép cặp theo nhãn, xem _pair_by_group_labels.
"""
import io
import os
import re
import unicodedata

MAGIC_OOXML = b"PK\x03\x04"
MAGIC_BIFF = b"\xd0\xcf\x11\xe0"

# Thứ tự quan trọng: nhóm CDPS phải xét TRƯỚC CDKT vì tên kiểu 'Bang can doi
# so phat sinh' cũng chứa cụm 'bang can doi'. Mỗi mẫu khớp trên tên file đã
# hạ chữ thường.
#
# PHÂN BIỆT SỐNG CÒN (đã xác nhận trên file thật 'bang can doi tai khoan
# 2016.xls' — bảng cân đối thử trên tài khoản 111, 112, 131... 911):
#   - 'bảng cân đối KẾ TOÁN'  = CDKT (báo cáo tài sản/nguồn vốn, mã 100-440);
#   - 'bảng cân đối TÀI KHOẢN' / 'bảng cân đối SỐ PHÁT SINH' = CDPS.
# Vì vậy mẫu 'bang can doi' TRẦN TRỤI không được suy ra CDKT nữa: một mình nó
# không phân biệt được hai loại trên. Chỉ 'can doi ke toan' (và chữ tắt
# 'cdkt') mới là CDKT.
_KIND_PATTERNS = (
    ("CDPS", r"cdsps|cdps|can doi so phat sinh|can doi phat sinh|can doi tai khoan"),
    ("CDKT", r"cdkt|can doi ke toan"),
    ("KQHDKD", r"kqkd|kqhdkd|ket qua"),
    ("LCTT", r"lctt|luu chuyen"),
)


def statement_kind(filename):
    """Suy ra loại báo cáo từ tên file. Trả về '' nếu không nhận ra."""
    name = os.path.basename(filename).lower()
    for kind, pat in _KIND_PATTERNS:
        if re.search(pat, name):
            return kind
    return ""


def sniff_format(path):
    """Dò định dạng thật theo magic bytes, KHÔNG tin đuôi file."""
    with open(path, "rb") as fh:
        head = fh.read(8)
    if head.startswith(MAGIC_OOXML):
        return "ooxml"
    if head.startswith(MAGIC_BIFF):
        return "biff"
    return "unknown"


def _rows_ooxml(path):
    import openpyxl
    with open(path, "rb") as fh:
        data = fh.read()
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    try:
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_biff(path):
    import xlrd
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)]
            for r in range(sh.nrows)]


def _read_rows(path):
    fmt = sniff_format(path)
    if fmt == "ooxml":
        return _rows_ooxml(path)
    if fmt == "biff":
        return _rows_biff(path)
    raise ValueError("Không nhận ra định dạng Excel: %s" % path)


def _to_int(v):
    """Chuyển ô Excel thành int. Trả về None nếu ô rỗng/không phải số.

    Dấu âm CHỈ được nhận theo 2 dạng: chuỗi bắt đầu bằng '-' (sau khi bỏ
    khoảng trắng hai đầu), hoặc bọc trong ngoặc đơn '(...)'. Dấu gạch ngang ở
    GIỮA chuỗi (vd. ngày tháng '2023-12-31', 'Ngày 30-06-2025') chỉ là dấu
    phân cách, không mang nghĩa âm và không được lẫn vào phần chữ số — độ lớn
    luôn được trích riêng từ các ký tự \\d.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    digits = re.sub(r"[^\d]", "", s)
    if not digits:
        return None
    val = int(digits)
    return -val if neg else val


_CODE_CELL = re.compile(r"^\s*(\d{1,3}[abc]?)\s*$")

# Dùng cho dự phòng hình học: ngưỡng tối thiểu để tin một cột là cột mã số /
# cột giá trị. Xem docstring của _find_columns_fallback.
_MIN_CODE_HITS = 5
_MIN_VALUE_ABS = 1000


def _row_code(row, code_col):
    """
    Mã chỉ tiêu tại cột code_col của một hàng, hoặc None nếu hàng không có mã.

    Đây là ĐỊNH NGHĨA DUY NHẤT về "hàng có mã": statement_detail() dùng nó để
    bóc giá trị, _numeric_hits() dùng nó để xác minh cột giá trị — hai nơi
    phải đồng nhất, nếu không bước xác minh sẽ đếm trên tập hàng khác với tập
    hàng được bóc.
    """
    if code_col >= len(row):
        return None
    raw = row[code_col]
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return str(int(round(raw)))
    m = _CODE_CELL.match(str(raw or ""))
    if m:
        return m.group(1)
    return None


def _numeric_hits(rows, hdr_row, code_col, col):
    """Số hàng CÓ MÃ (dưới hàng tiêu đề) mà ô tại cột `col` là số.

    Chỉ đếm trên hàng có mã để các ô số lạc ngoài bảng (ngày tháng ở dòng ký
    tên 'Lập, ngày... năm 2023', số hiệu thông tư ở đầu trang...) không làm
    một cột rỗng trông như có dữ liệu.
    """
    n = 0
    for row in rows[hdr_row + 1:]:
        if _row_code(row, code_col) is None:
            continue
        if col < len(row) and _to_int(row[col]) is not None:
            n += 1
    return n


# Ngưỡng cho _nudge_empty_value_col, tính theo TỶ LỆ trên số hàng có mã (số
# tuyệt đối không dùng được: bảng CDKT 115 hàng và bảng KQKD 20 hàng cùng một
# dạng lỗi). "Rỗng" không có nghĩa là 0 ô tuyệt đối: 'CDKT Q2.xls' có đúng 1 ô
# số lạc trong cột marker (hàng '270', c9=248197922) trên 115 hàng có mã —
# vẫn phải coi là rỗng. Cột thay thế thì ngược lại phải DÀY thật sự: các bản
# in phần mềm kế toán thuộc gia đình lỗi này in số (kể cả 0) trên gần như mọi
# hàng, nên cột kề thưa (vd. Thuyết minh 'VI.25' -> _to_int ra 25) không đạt
# ngưỡng dày và không bao giờ được chọn thay.
_EMPTY_COL_FRAC = 0.05   # <= 5% hàng có mã có ô số -> coi như cột rỗng
_DENSE_COL_FRAC = 0.5    # >= 50% hàng có mã có ô số -> mới đáng tin để thay


def _nudge_empty_value_col(rows, hdr_row, code_col, col, taken):
    """
    Xác minh cột giá trị `col` neo theo hàng đánh số; trả về cột nên dùng.

    Cạm bẫy 5 (đầu file): hàng đánh số nằm trên ô merge nên marker '4'/'5' có
    thể lệch 1 cột so với cột giá trị thật — cột được neo khi đó RỖNG (theo
    _EMPTY_COL_FRAC) trên các hàng có mã. Cột không rỗng thì giữ nguyên
    (không đụng vào file lành). Cột rỗng thì dò đúng 2 cột kề (lệch-1 do
    merge, không dò xa hơn), bỏ qua cột mã số và các cột đã bị chiếm
    (`taken`), chỉ nhận cột kề DÀY thật sự (theo _DENSE_COL_FRAC), chọn cột
    nhiều ô số hơn; hoà thì ưu tiên TRÁI — cả ba file lỗi đã xác nhận
    (KQKD_2023, cdkt_6t, CDKT Q2) đều có marker nằm bên PHẢI cột giá trị
    thật. Không cột kề nào đạt thì giữ nguyên `col` (không đoán xa).
    """
    n_coded = sum(1 for row in rows[hdr_row + 1:]
                  if _row_code(row, code_col) is not None)
    if _numeric_hits(rows, hdr_row, code_col, col) > n_coded * _EMPTY_COL_FRAC:
        return col
    best, best_hits = col, 0
    for cand in (col - 1, col + 1):
        if cand < 0 or cand == code_col or cand in taken:
            continue
        hits = _numeric_hits(rows, hdr_row, code_col, cand)
        if hits >= max(n_coded * _DENSE_COL_FRAC, 1) and hits > best_hits:
            best, best_hits = cand, hits
    return best


def _find_columns_detail(rows):
    """
    Dò đầy đủ nguồn gốc: (hàng tiêu đề, cột mã số, cột năm nay, cột năm
    trước, strategy, column_order) hoặc None nếu không xác định được.

      - strategy = "numbering": neo vào hàng đánh số thứ tự cột
        ('1','2','3','4','5') vì nhãn tiêu đề thường là mojibake nên không so
        khớp chữ được. Hàng đó nằm ngay dưới hàng tiêu đề và có ít nhất 4 ô
        là số nguyên nhỏ tăng dần bắt đầu từ 1. column_order LUÔN là
        "positional": đường này không đọc nhãn chữ, cột năm nay/năm trước
        suy ra từ VỊ TRÍ trong dãy số thứ tự (quy ước cố định 4 = năm nay,
        5 = năm trước), không phải từ nhãn tiêu đề.
      - strategy = "fallback": không dò được hàng đánh số (layout khác, phần
        mềm kế toán khác) -> dự phòng thuần hình học ở _find_columns_fallback,
        column_order do hàm đó quyết định ("labels" nếu nhãn tiêu đề còn đọc
        được và xác định rõ năm nay/năm trước, "positional" nếu không).

    Đây là nơi DUY NHẤT cài đặt logic dò cột — cả _find_columns() (wrapper
    tương thích ngược, chỉ trả 4 giá trị đầu) lẫn statement_detail() đều gọi
    qua hàm này, tránh hai đường logic tách rời dễ lệch nhau.
    """
    for r, row in enumerate(rows[:25]):
        nums = [(c, _to_int(v)) for c, v in enumerate(row) if _to_int(v) is not None]
        seq = [(c, n) for c, n in nums if 1 <= n <= 9]
        if len(seq) >= 4 and [n for _, n in seq][:4] == [1, 2, 3, 4]:
            cols = [c for c, _ in seq]
            # cột 1 = Chỉ tiêu, 2 = Mã số, 3 = Thuyết minh, 4 = Năm nay, 5 = Năm trước
            code_col = cols[1]
            cur_col = cols[3]
            prior_col = cols[4] if len(cols) >= 5 else cols[3] + 1
            # Cạm bẫy 5: marker '4'/'5' trên ô merge có thể lệch 1 cột so với
            # cột giá trị thật -> xác minh từng cột, cột rỗng thì dò cột kề.
            # Chỉnh cur trước rồi mới tới prior; cột đối phương (theo trạng
            # thái mới nhất) bị loại khỏi ứng viên để hai cột không trùng nhau.
            cur_col = _nudge_empty_value_col(rows, r, code_col, cur_col,
                                             (prior_col,))
            prior_col = _nudge_empty_value_col(rows, r, code_col, prior_col,
                                               (cur_col,))
            return r, code_col, cur_col, prior_col, "numbering", "positional"

    fb = _find_columns_fallback(rows)
    if fb is None:
        return None
    hdr_row, code_col, cur_col, prior_col, column_order = fb
    return hdr_row, code_col, cur_col, prior_col, "fallback", column_order


def _find_columns(rows):
    """
    Dò (hàng tiêu đề, cột mã số, cột năm nay, cột năm trước).

    Wrapper tương thích ngược quanh _find_columns_detail(): giữ nguyên chữ ký
    4 giá trị trả về như trước đây (không có strategy/column_order) để không
    phá vỡ các chỗ gọi hiện có. Cần biết cột năm nay/năm trước được xác định
    bằng nhãn hay theo vị trí mặc định thì gọi statement_detail() thay vì
    hàm này.
    """
    found = _find_columns_detail(rows)
    if found is None:
        return None
    hdr_row, code_col, cur_col, prior_col, _strategy, _column_order = found
    return hdr_row, code_col, cur_col, prior_col


def _is_code_cell(v):
    """
    True nếu ô trông như mã chỉ tiêu: số nguyên nhỏ 1-999, hoặc chuỗi
    '132'/'132a' (khác 0).

    Giới hạn 1-999 cho nhánh int CỐ Ý khớp với \\d{1,3} của nhánh chuỗi (mã
    số BCTC luôn >= 1, không có "mã số 0"), vì hai lý do đã xác nhận thật
    trên corpus:
      - Không chặn biên trên: một ô tiền tệ nguyên LỚN (vd. 304006836) đọc
        bằng openpyxl trả về kiểu int trần (không phải float) sẽ bị đếm
        nhầm thành "mã chỉ tiêu" và cột giá trị có thể lấn cột mã số
        (CDSPS.XLS).
      - Không loại trừ 0: trong bảng cân đối phát sinh, cột Nợ/Có thường
        bằng 0 ở phần lớn các dòng (không phát sinh) — nếu tính cả 0, một
        cột giá trị "phần lớn bằng 0" có thể có SỐ LƯỢNG ô khớp nhiều hơn cả
        cột mã số thật, đè luôn cột mã số (cũng xác nhận trên CDSPS.XLS).
    """
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return 1 <= v <= 999
    if isinstance(v, str):
        m = _CODE_CELL.match(v)
        if not m:
            return False
        digits = m.group(1).rstrip("abc")
        return digits != "" and int(digits) >= 1
    return False


# ----------------------------------------------------------------------
# Dò thứ tự cột năm nay/năm trước qua NHÃN tiêu đề, khi nhãn còn đọc được.
#
# Rất nhiều file có nhãn mojibake TCVN3 ('M· sè' thay vì 'Mã số') nên đây
# KHÔNG THỂ là chiến lược chính — chỉ là một ưu tiên áp dụng TRƯỚC khi rơi về
# mặc định vị trí (trái = năm nay) trong _find_columns_fallback. Bản bỏ dấu
# cục bộ dưới đây CỐ Ý không import từ bctc/parser.py (nơi có strip_accents/
# norm gần như y hệt): tests/ phải độc lập với mã nguồn production của bộ
# đọc "đáp án" này.
# ----------------------------------------------------------------------
def _strip_accents_local(s):
    s = s.replace("đ", "d").replace("Đ", "D")
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm_local(s):
    return re.sub(r"\s+", " ", _strip_accents_local(s).lower()).strip()


# Đã xác minh: nhãn mojibake (vd. 'N¨m nay' đọc lệch từ TCVN3) KHÔNG chuẩn
# hoá trùng với các mốc dưới đây ('¨' giải mã NFKD ra khoảng trắng + dấu kết
# hợp, chứ không ra 'a') nên mojibake không tự khớp nhầm thành current/prior.
# Điều đó KHÔNG có nghĩa mọi cụm tiếng Việt hợp lệ chỉ khớp một phía: nhiều
# cụm tự nhiên (vd. "Số dư cuối năm trước", "Cuối kỳ trước") chứa cả mốc năm
# nay ("cuối năm") LẪN mốc năm trước ("năm trước") làm con của cùng một cụm.
# _year_marker() xử lý ca đó bằng cách trả None khi khớp CẢ HAI danh sách —
# nhập nhằng bị bỏ qua có chủ đích, KHÔNG được đoán đại một phía; bên gọi
# (_order_by_header_labels) sẽ không dùng ô đó làm căn cứ và rơi về mặc định
# vị trí (trái = năm nay) thay vì gán liều một kết quả sai mà tự tin.
_CUR_YEAR_MARKERS = ("nam nay", "ky nay", "so cuoi nam", "cuoi ky", "cuoi nam")
_PRIOR_YEAR_MARKERS = ("nam truoc", "ky truoc", "so dau nam", "dau ky", "dau nam")
_LABEL_SCAN_ABOVE = 3  # số hàng phía TRÊN hàng tiêu đề ứng viên cũng được xét


def _year_marker(cell):
    """'current' / 'prior' / None — ô này có khớp nhãn năm nay/năm trước không.

    Khớp CẢ HAI danh sách (vd. "Số dư cuối năm trước" vừa chứa "cuối năm" vừa
    chứa "năm trước") nghĩa là nhập nhằng: trả về None, KHÔNG đoán đại một
    phía. Không khớp danh sách nào, ô rỗng, hoặc không phải chuỗi cũng trả về
    None — chỉ trả 'current'/'prior' khi khớp ĐÚNG MỘT phía.
    """
    if not isinstance(cell, str):
        return None
    text = _norm_local(cell)
    if not text:
        return None
    is_current = any(m in text for m in _CUR_YEAR_MARKERS)
    is_prior = any(m in text for m in _PRIOR_YEAR_MARKERS)
    if is_current and is_prior:
        return None
    if is_current:
        return "current"
    if is_prior:
        return "prior"
    return None


def _order_by_header_labels(rows, hdr_row, col_a, col_b):
    """
    Xét hàng tiêu đề ứng viên (hdr_row) và vài hàng phía trên nó (xem
    _LABEL_SCAN_ABOVE), tìm nhãn năm nay/năm trước tại 2 cột giá trị ứng viên
    (col_a, col_b — thứ tự bất kỳ, không giả định cột nào là năm nay).

    Nếu ĐÚNG MỘT trong hai cột khớp nhãn năm nay và cột còn lại khớp nhãn năm
    trước, trả về (cột năm nay, cột năm trước) theo đúng nhãn. Nếu ở mọi hàng
    trong cửa sổ quét đều không tìm được cặp khớp rõ ràng như vậy (nhãn
    mojibake, thiếu nhãn, hoặc cả hai cột cùng khớp một loại mốc), trả về
    None để bên gọi tự quyết theo mặc định vị trí (trái = năm nay).
    """
    lo = max(hdr_row - _LABEL_SCAN_ABOVE, 0)
    for r in range(hdr_row, lo - 1, -1):
        row = rows[r]
        cell_a = row[col_a] if col_a < len(row) else None
        cell_b = row[col_b] if col_b < len(row) else None
        marker_a = _year_marker(cell_a)
        marker_b = _year_marker(cell_b)
        if marker_a == "current" and marker_b == "prior":
            return col_a, col_b
        if marker_a == "prior" and marker_b == "current":
            return col_b, col_a
    return None


def _column_year_marker(rows, hdr_row, col):
    """Nhãn năm nay/năm trước của MỘT cột trong cửa sổ tiêu đề, hoặc None.

    Quét từ hdr_row ngược lên (cùng cửa sổ _LABEL_SCAN_ABOVE với
    _order_by_header_labels), lấy nhãn ở hàng GẦN dữ liệu nhất: với bảng
    2 tầng header ('Trong kỳ' đè trên 'Năm nay'/'Năm Trước'), tầng dưới mới
    nói cột này là năm nào — tầng trên ('Lũy kế từ ĐẦU NĂM' chứa mốc 'đầu
    năm') mà thắng thì cột năm-nay của nhóm Lũy kế bị dán nhầm nhãn prior.
    """
    lo = max(hdr_row - _LABEL_SCAN_ABOVE, 0)
    for r in range(hdr_row, lo - 1, -1):
        row = rows[r]
        marker = _year_marker(row[col] if col < len(row) else None)
        if marker is not None:
            return marker
    return None


def _pair_by_group_labels(rows, hdr_row, value_cols, value_counts):
    """
    Ghép cặp (cột năm nay, cột năm trước) THEO NHÃN giữa các cột ứng viên,
    dùng cho bảng có nhiều hơn 2 cột giá trị chia NHÓM header (cạm bẫy 6 —
    'BTHT ..._KQKD.xls': Trong kỳ nay/trước + Lũy kế nay/trước; 2 cột nhiều
    số nhất là 2 cột Năm-Trước của 2 nhóm khác nhau).

    Chỉ quyết định khi đọc được CẢ HAI phía trong các ứng viên: cột năm nay =
    cột mang nhãn năm-nay có nhiều ô số nhất (hoà thì lấy trái nhất); cột năm
    trước = cột mang nhãn năm-trước GẦN cột năm nay nhất — cùng nhóm header
    thì đứng cạnh nhau, nên "gần nhất" chính là "cùng nhóm" mà không phải
    đoán ranh giới nhóm từ ô merge. Thiếu một phía (nhãn mojibake/thiếu/nhập
    nhằng) thì trả về None để bên gọi giữ nguyên đường cũ.
    """
    markers = {c: _column_year_marker(rows, hdr_row, c) for c in value_cols}
    cur_cands = [c for c in value_cols if markers[c] == "current"]
    prior_cands = [c for c in value_cols if markers[c] == "prior"]
    if not cur_cands or not prior_cands:
        return None
    cur_col = max(cur_cands, key=lambda c: (value_counts[c], -c))
    prior_col = min(prior_cands, key=lambda c: (abs(c - cur_col), c))
    return cur_col, prior_col


def _find_columns_fallback(rows):
    """
    Dự phòng khi không có hàng đánh số cột — chỉ dùng khi chiến lược chính ở
    _find_columns thất bại. Không so khớp chữ để CHỌN cột (nhãn có thể là
    mojibake như 'M· sè' thay vì 'Mã số'), chỉ dựa vào HÌNH DẠNG ô:

      1. Cột mã số: cột có nhiều ô khớp _is_code_cell nhất trên toàn bộ
         bảng. Cần >= _MIN_CODE_HITS ô mới coi là đáng tin, nếu không thất
         bại (trả về None, bên gọi tự raise ValueError).
      2. Cột giá trị: trong các cột nằm BÊN PHẢI cột mã số, đếm ô số có
         |giá trị| >= _MIN_VALUE_ABS; mọi cột có ít nhất 1 ô như vậy là ỨNG
         VIÊN. Cột "Thuyết minh" (vd. 'VI.25') thường nằm xen giữa nhưng
         không phải số lớn nên không lẫn vào đây — ngưỡng độ lớn mới là chốt
         chặn thật, không dựa vào việc đoán đúng vị trí cột Thuyết minh. Nếu
         có dưới 2 ứng viên, thất bại.
      3. Hàng tiêu đề: hàng ngay TRƯỚC hàng đầu tiên chứa mã số đã dò được
         (không âm), để read_statement bắt đầu quét đúng chỗ.
      4. Chọn cặp + thứ tự (năm nay, năm trước): nếu nhãn tiêu đề đọc được
         CẢ hai phía trong các ứng viên thì ghép cặp theo nhãn
         (_pair_by_group_labels) — nhãn LUÔN thắng vị trí vì đây là căn cứ
         chắc chắn nhất khi có, và với bảng >2 cột chia nhóm header (cạm bẫy
         6) "2 cột nhiều số nhất" có thể là 2 cột Năm-Trước xuyên nhóm. Nếu
         nhãn không đủ (mojibake / thiếu / mơ hồ): lấy 2 cột nhiều ô số
         nhất, thử _order_by_header_labels trên đúng 2 cột đó, còn không thì
         GIỮ mặc định đã dùng từ trước: cột trái = năm nay, phải = năm trước.

    Trả về (hàng tiêu đề, cột mã số, cột năm nay, cột năm trước,
    column_order) — column_order là "labels" hoặc "positional", cho biết
    bước 4 ở trên quyết định thứ tự bằng cách nào. Trả về None nếu bước 1
    hoặc 2 thất bại.
    """
    ncols = 0
    for row in rows:
        ncols = max(ncols, len(row))
    if ncols == 0:
        return None

    code_counts = [0] * ncols
    first_code_row = [None] * ncols
    for r, row in enumerate(rows):
        for c, v in enumerate(row):
            if _is_code_cell(v):
                code_counts[c] += 1
                if first_code_row[c] is None:
                    first_code_row[c] = r

    code_col = max(range(ncols), key=lambda c: code_counts[c])
    if code_counts[code_col] < _MIN_CODE_HITS:
        return None

    value_counts = [0] * ncols
    for row in rows:
        for c in range(code_col + 1, len(row)):
            n = _to_int(row[c])
            if n is not None and abs(n) >= _MIN_VALUE_ABS:
                value_counts[c] += 1

    value_cols = [c for c in range(code_col + 1, ncols) if value_counts[c] > 0]
    if len(value_cols) < 2:
        return None

    hdr_row = max(first_code_row[code_col] - 1, 0)

    # Cạm bẫy 6: khi nhãn đọc được cả 2 phía, ghép cặp theo nhãn TRƯỚC khi
    # đếm số — "2 cột nhiều số nhất" có thể là 2 cột Năm-Trước xuyên nhóm.
    # Với bảng 2 cột thường gặp, đường này chọn đúng 2 cột đó theo đúng thứ
    # tự nhãn — trùng kết quả đường cũ (top-2 + _order_by_header_labels).
    paired = _pair_by_group_labels(rows, hdr_row, value_cols, value_counts)
    if paired is not None:
        cur_col, prior_col = paired
        return hdr_row, code_col, cur_col, prior_col, "labels"

    value_cols.sort(key=lambda c: value_counts[c], reverse=True)
    col_a, col_b = value_cols[:2]

    labeled = _order_by_header_labels(rows, hdr_row, col_a, col_b)
    if labeled is not None:
        cur_col, prior_col = labeled
        column_order = "labels"
    else:
        cur_col, prior_col = sorted((col_a, col_b))
        column_order = "positional"

    return hdr_row, code_col, cur_col, prior_col, column_order


def statement_detail(path):
    """
    Đọc file đáp án, trả về CẢ số liệu lẫn nguồn gốc cách dò cột năm nay/năm
    trước — để bên gọi (regression harness) phân biệt được "đọc sai" (lỗi
    thật của bộ đọc) với "đo sai" (lỗi phép đo, vd. 2 năm bị hoán đổi mà
    không có tín hiệu nào báo). Trả về dict:

      "values"       : {ma_so: (nam_nay, nam_truoc)} — giống hệt những gì
                        read_statement() trả về (mã số giữ nguyên dạng in
                        trong file, '1' chứ không phải '01'; chuẩn hoá là
                        trách nhiệm của bên so sánh).
      "strategy"     : "numbering" nếu neo được vào hàng đánh số cột (chiến
                        lược chính), "fallback" nếu phải dò hình học.
      "column_order" : "labels" nếu nhãn tiêu đề (vd. 'Kỳ này'/'Kỳ trước')
                        còn đọc được và quyết định thứ tự năm nay/năm trước;
                        "positional" nếu dùng mặc định trái = năm nay (nhãn
                        mojibake, không có nhãn phù hợp, hoặc đi theo hàng
                        đánh số).

    Raise ValueError nếu không dò được cột (không âm thầm trả về rỗng).
    """
    rows = _read_rows(path)
    found = _find_columns_detail(rows)
    if not found:
        raise ValueError("Không dò được hàng tiêu đề trong %s" % path)
    hdr_row, code_col, cur_col, prior_col, strategy, column_order = found

    values = {}
    for row in rows[hdr_row + 1:]:
        code = _row_code(row, code_col)
        if not code:
            continue
        cur = _to_int(row[cur_col]) if cur_col < len(row) else None
        prior = _to_int(row[prior_col]) if prior_col < len(row) else None
        values[code] = (cur, prior)

    return {"values": values, "strategy": strategy, "column_order": column_order}


def read_statement(path):
    """
    Trả về {ma_so: (nam_nay, nam_truoc)}.

    Mã số giữ nguyên dạng in trong file ('1' chứ không phải '01') — việc chuẩn
    hoá là trách nhiệm của bên so sánh. Wrapper mỏng quanh statement_detail()
    — dùng hàm đó thay vì hàm này nếu cần biết cột năm nay/năm trước được
    xác định bằng nhãn tiêu đề hay theo vị trí mặc định.
    """
    return statement_detail(path)["values"]
